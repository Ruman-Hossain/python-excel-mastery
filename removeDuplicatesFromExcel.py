import openpyxl

print("opening workbook")
wb = openpyxl.load_workbook("duplicates.xlsx")
print("Excel Workbook Opened...")
sheet = wb['NB Permits']
print(str(sheet)+" Reading...")
applicantList=[]
mailList=[]
ownerList=[]
ownerMailList=[]
for r in range(2,1001):

	print("========================= %d =============================="%r)

	AppName = sheet.cell(row=r, column=10).value
	AppMail = sheet.cell(row=r, column=11).value
	OwnName = sheet.cell(row=r, column=14).value
	Onwmail = sheet.cell(row=r, column=15).value

	if AppName in applicantList:
		print("%s already Exist" %AppName)
		sheet.cell(row=r, column=10).value=None
	if AppName not in applicantList:
		applicantList.append(AppName)
		print(applicantList[-1])

	if AppMail in mailList:
		print("%s already Exist" %AppMail)
		sheet.cell(row=r, column=11).value=None
	if AppMail not in mailList:
		mailList.append(AppMail)
		print(mailList[-1])

	if OwnName in ownerList:
		print("%s already Exist" %OwnName)
		sheet.cell(row=r, column=14).value=None
	if OwnName not in ownerList:
		ownerList.append(OwnName)
		print(ownerList[-1])

	if Onwmail in ownerMailList:
		print("%s already Exist" %Onwmail)
		sheet.cell(row=r, column=15).value=None
	if Onwmail not in ownerMailList:
		ownerMailList.append(Onwmail)
		print(ownerMailList[-1])


print("Final saving")
wb.save('duplicates.xlsx')
