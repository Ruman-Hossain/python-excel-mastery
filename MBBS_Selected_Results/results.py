from selenium import webdriver
from selenium.webdriver.chrome.options import Options as ChromeOptions
import openpyxl
import time
import random

driver=webdriver.Chrome('../chromedriver')
wb=openpyxl.load_workbook("mbbsResultData.xlsx")
print("Excel Workbook Opened...")
sheet=wb['result']
print(str(sheet)+" Reading...")
for r in range(256,4353):
    # READ EXCEL SHEET FOR ROLL NUMBER
    student_roll = sheet.cell(row=r,column=3).value
    link=f'https://result.dghs.gov.bd/mbbs/'
    # SEARCH RESULT IN GIVEN WEBSITE URL
    driver.get(link)
    driver.find_element_by_name('roll').send_keys(student_roll)
    driver.find_element_by_name('Result').click()
    # CATCH COMPLETE RESULT INFORMATION
    roll = driver.find_element_by_xpath('//*[@id="rockartists"]/tbody/tr[1]/td').text
    name = driver.find_element_by_xpath('//*[@id="rockartists"]/tbody/tr[2]/td').text
    test_score = driver.find_element_by_xpath('//*[@id="rockartists"]/tbody/tr[3]/td').text
    merit_score = driver.find_element_by_xpath('//*[@id="rockartists"]/tbody/tr[4]/td').text
    merit_position = driver.find_element_by_xpath('//*[@id="rockartists"]/tbody/tr[5]/td').text
    alloted_college = driver.find_element_by_xpath('//*[@id="rockartists"]/tbody/tr[6]/td').text
    status = driver.find_element_by_xpath('//*[@id="rockartists"]/tbody/tr[7]/td').text
    # WRITE OUTPUT TO EXCEL FILE AS ACTUAL SCORE
    sheet.cell(row=r,column=7).value= merit_position
    sheet.cell(row=r,column=8).value= name
    sheet.cell(row=r,column=9).value= roll
    sheet.cell(row=r,column=10).value= test_score
    sheet.cell(row=r,column=11).value= merit_score
    sheet.cell(row=r,column=12).value= alloted_college
    sheet.cell(row=r,column=13).value= status
    # DATA PREVIEW IN TERMINAL
    print("ROLL           : " + roll)
    print("Name           : " + name)
    print("Test Score     : " + test_score)
    print("Merit Score    : " + merit_score)
    print("Merit Position : " + merit_position )
    print("Alloted College: " + alloted_college )
    print("Status         : " + status )
    wb.save('mbbsResultData.xlsx')
    print("DATA INSERTED SUCCESSFULLY AT ROW : ",r)
    # EXIT
driver.quit()
