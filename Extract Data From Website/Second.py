from selenium import webdriver
from selenium.webdriver.chrome.options import Options as ChromeOptions
import openpyxl
import time
import random

import requests
from urllib.request import urlopen
from lxml import etree
import os

import loopCookies
#chrome_options = ChromeOptions()
# PROXY = "18.223.1.180:80" # IP:PORT or HOST:PORT
# chrome_options = webdriver.ChromeOptions()
# chrome_options.add_argument('--proxy-server=http://%s' % PROXY)
# driver = webdriver.Chrome('Drivers/chromedriver', options=chrome_options)

print("opening workbook")
wb = openpyxl.load_workbook("userdata.xlsx")
print("Excel Workbook Opened...")
sheet = wb['NB Permits']
print(str(sheet)+" Reading...")
rowsNotProcessed = open("error on rows.txt", "w")
cnt = 1
# cookies = {"__utma": "24711658.548814779.1599230964.1599386945.1599395006.7", "__utmb": "24711658.20.10.1599395006", "__utmc": "24711658", "__utmt": "1", "__utmz": "24711658.1599230964.1.1.utmcsr=(direct)|utmccn=(direct)|utmcmd=(none)", "_abck": "D6A74F5C2A3FE944373600C5FAA892CF~0~YAAQ5PQPZy5eIVF0AQAATG+JYwQUaK8e+7TC6FNqBNblUPqjqdBYv3XBbDnprvpj48vsgXUScMs2A1Ey62hDezsyCXtnI1C46SODXA2sH2BiRQnxE40uzzPeGy1ouh9tv8EQY+9shfqFaQPxn3DoI4dgOBA97SWYqUQpZtZRwsx8Yba/A42Ll6XL0/D0qF5nUYJUBCuyDwKPY0a+U9PTDnqF9UyP4rvIeQzdPNXrsJt24Hc95qu0BqBvy1t1L90CGy/cCirxjB7DMUxi098WNeyqzYXdN4n2EsvCEk6Gow==~-1~1-jhoYNZdSWZ-10000-10-1000-1||-1||~-1", "_ga": "GA1.2.1268737468.1599383394", "ak_bmsc": "22D85C193B09678675D76BEE3B62B3DB670FF4E4C625000044D8545FEB5A0F6F~plcFYInPGsYE+3hslMWEVf3IfdbXbEnB5sk6wH5z3DjZ8ukaf73eWrUJQp5aRlKHEyoClFHkReYh4CQOeoeb9hJ3o96T3V9MPJItgCEnuufPJfUhKBb886ZQ50BsSk3c3jD4r6u83AAXpIrSal7jZstZgRQvEQ7020CFKtnWwxrpUAZv85hjwApAP+pcBgJxpDDaBCWMyCSfsp5visWu2sGtbMZ0jWFvymrp60P8zAk0r1+ivVNm1qZJzRR4U/HAq7", "bm_sv": "9C8D9FD1329E2F375B0D3748835F8979~dU1DYjSER9uEqcDQWIN6yJNI0ziulHrWFog2hV8oYTsICRRRNkKCrG6/M5K7eADp5b7axIbXxfMU8JG1osj5vu7SD2mLa16V/uYwF9j8pByzr6ZN6afGz6t0tCQFgTDPWID2+XPOmuOKpJEZXE7Z9Q==", "bm_sz": "E7FB38A0F7ED82196243FC708C883812~YAAQ5PQPZxFeIVF0AQAAwsxsYwk1SFKxz1T9TSpk9lIhC1jytciu21QuafKpmawfgR4CZxPCDqUl+mzl4PoIDlvkmE4Np2rPLG7rsGdcKf8vaPTDs4UpS61XBt00Ao3Xnz/MnhS4k1gs0aZAQlkduXnJdLziMLYhn5P3xi4Njr0UV/Vj64Hyee7vqGU9", "aa": 
# "45896", "JSESSIONID": "951896CFBC1972AB79522BB70DF7DDFC"} 
cookies=loopCookies.cookies
print(cookies)
headers = {
	"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.83 Safari/537.36",
	"Accept-Language":"en-US,en;q=0.9,bn;q=0.8",
	"Accept-Encoding":"gzip, deflate, br",
	"Accept":"text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
	"Content-Type":"text/html"
}
looper = loopCookies.iterate
print(looper)
for r in range(looper,looper+200):
    try:
        jobid = sheet.cell(row=r, column=1).value

        url = "http://a810-bisweb.nyc.gov/bisweb/JobsQueryByNumberServlet?passjobnumber=%d&passdocnumber=&go10=+GO+&requestid=0"%jobid
        print("for url:", url)

        print("Input Given in the From ID: "+str(jobid))
        print("Wait Data Loading...")

        response = requests.get(url=url, cookies=cookies, headers=headers)
        html = response.text
		
        f = open("aa.html", "w")
        f.write(html)
        local = "file:///" + os.path.abspath(f.name)
        f.close()
        response = urlopen(local)
        htmlparser = etree.HTMLParser()
        tree = etree.parse(response, htmlparser)

        if tree.xpath('/html/head/title/text()')[0] != "Application Details":
            print("NB Permits UNREACHED LEFT BLANK ROW "+str(r))
            print(r, "UNREACHED", file=rowsNotProcessed)
            sheet.cell(row=r, column=18).value = "Try Finding Again"
            print("Access denied at %d"%r)
            break
        else:
            print("NB Permits DATA FOUND")
            applicantname = tree.xpath(
                "/html/body/center/table[8]/tr[3]/td[2]/text()")[0]
            applicantmail = tree.xpath(
                "/html/body/center/table[8]/tr[6]/td[2]/text()")[0]

            ownername = tree.xpath(
                "/html/body/center/table[35]/tr[3]/td[2]/text()")[0]
            ownermail = tree.xpath(
                "/html/body/center/table[35]/tr[7]/td[2]/text()")[0]

            sheet.cell(row=r, column=10).value = applicantname
            sheet.cell(row=r, column=11).value = applicantmail
            sheet.cell(row=r, column=14).value = ownername
            sheet.cell(row=r, column=15).value = ownermail
            # wb.save('userdata.xlsx')
            print("DATA INSERTED SUCCESSFULLY IN ROW "+str(r))
            print("%s %s %s %s"%(applicantname, applicantmail, ownername, ownermail))
            time.sleep(2+random.randint(0, 3))
    except:
        print(r, "CRUSHED", file=rowsNotProcessed)
    finally:
        if cnt % 10 == 0:
            rowsNotProcessed.close()
            wb.save('userdata.xlsx')
            wb = openpyxl.load_workbook("userdata.xlsx")
            print("Excel Workbook Opened...")
            sheet = wb['NB Permits']
            print(str(sheet)+" Reading...")
            rowsNotProcessed = open("error on rows.txt", "a")
    cnt += 1
    time.sleep(5+random.randint(0, 3))
rowsNotProcessed.close()
print("Final saving")
wb.save('userdata.xlsx')
