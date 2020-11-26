from selenium import webdriver
from selenium.webdriver.chrome.options import Options as ChromeOptions
import openpyxl
import time
import random

for r in range(70,185):
   wb=openpyxl.load_workbook("CarpenterData.xlsx")
   sheet=wb['carpenter']
   print(str(sheet)+" Reading Row :"+str(r))
   try:
      address=sheet.cell(row=r,column=5).value
      address=address.replace('\n',',')
      sheet.cell(row=r,column=5).value=address
   except:
      print(str(r) +" is Empty")
   wb.save('CarpenterData.xlsx')
   print(address)
   