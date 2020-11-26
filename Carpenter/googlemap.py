from selenium import webdriver
from selenium.webdriver.chrome.options import Options as ChromeOptions
import openpyxl
import time
import random
def find_gmap(r):
	wb=openpyxl.load_workbook("CarpenterData.xlsx")
	print("Excel Workbook Opened...")
	sheet=wb['carpenter']

	url ='https://www.google.com/maps/@25.7473239,89.2477309,14z'
	driver=webdriver.Chrome('../Drivers/chromedriver')
	driver.get(url)

	company = sheet.cell(row=r,column=2).value
	address=sheet.cell(row=r,column=5).value
	address=address.replace('\n',',')
	search =company+','+address
	print(search)
	driver.find_element_by_xpath('/html/body/jsl/div[3]/div[9]/div[3]/div[1]/div[1]/div[1]/div[2]/form/div/div[3]/div/input[1]').send_keys(search)
	driver.find_element_by_id('searchbox-searchbutton').click()
	print("Wait Map Loading...")
	sheet.cell(row=r,column=5).value=address
	wb.save('CarpenterData.xlsx')
	print("DATA INSERTED SUCCESSFULLY IN ROW "+str(r))
	driver.quit()
for r in range(2,194):
	find_gmap(r)