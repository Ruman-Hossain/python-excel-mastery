from selenium import webdriver
from selenium.webdriver.chrome.options import Options as ChromeOptions
import openpyxl
import time
import random

wb=openpyxl.load_workbook("oasisnyc.xlsx")
print("Excel Workbook Opened...")
sheet=wb['tbl_property']
print(str(sheet)+" Reading...")

for r in range(2612,2613):

  print("Searching Data of Row : %s"%r)
  driver=webdriver.Chrome('../chromedriver')


  address=sheet.cell(row=r,column=4).value
  borough=sheet.cell(row=r,column=5).value
  print('---------------------INPUT FORM EXCEL---------------------')
  print('Address : '+address)
  print('Borough : '+borough)

  link=f'http://www.oasisnyc.net/map.aspx'
  print(link)
  driver.get(link)

  driver.find_element_by_id('ext-comp-1065__locationReport').click()
  driver.implicitly_wait(30)
  #------------------------------INPUT ADDRESS AND BOROUGH--------------------------#
  borough.replace(" ","")
  driver.find_element_by_name('address').send_keys(address)
  driver.find_element_by_name('addrBorough').send_keys(borough)
  driver.implicitly_wait(10)
  driver.find_element_by_id('ext-gen118').click()
  print('#-------------------------EXTRACTED DATA -------------------------#')
  # sheet.cell(row=r,column=6).value=zoning
  # sheet.cell(row=r,column=7).value=lotSize
  # sheet.cell(row=r,column=9).value=buildingArea

  # print("Primary Zoning : "+zoning)
  # print("Lot Size : "+lotSize)
  # print("Building Area : "+buildingArea)
  # wb.save('CarpenterData.xlsx')
  # print("DATA INSERTED SUCCESSFULLY IN ROW "+str(r))
