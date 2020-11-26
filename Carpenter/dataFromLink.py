from selenium import webdriver
from selenium.webdriver.chrome.options import Options as ChromeOptions
import openpyxl
import time
import random

wb=openpyxl.load_workbook("CarpenterData.xlsx")
print("Excel Workbook Opened...")
sheet=wb['carpenter']
print(str(sheet)+" Reading...")
for r in range(4,5):
  print("Searching Data of Row : %s"%r)
  driver=webdriver.Chrome('../Drivers/chromedriver')
  link=sheet.cell(row=r,column=1).value
  link=link.strip()
  print(link)
  driver.get(link)
  #------------------------------BUSINESS NAME--------------------------#
  try:
    business = driver.find_element_by_css_selector('body > div.main-view > div > div > bdp-header > div > div.text-box > div > div.headline-holder > h1').text
  except:
    business = ''
  #--------------------------------PHONE NUMBER ----------------------#
  try:
    phone = driver.find_element_by_css_selector('#phone-1 > bdp-phone > span > span').text
    phone=phone.replace(' ','')
    phone=phone.replace('(','')
    phone=phone.replace('-','')
    phone=phone.replace(')','')
  except:
    phone = ''
  #---------------------------------ADDRESS---------------------------#
  try:
    address = driver.find_element_by_css_selector('#address > div').text
    address=address.replace('\n',',')
  except:
    address = ''
  #---------------------------------TITLE-----------------------------#
  try:
    title = driver.find_element_by_css_selector('#details-business-title').text
  except:
    title = ''
  #-------------------------------TAGLINE----------------------------#
  try:
    tagline = driver.find_element_by_css_selector('#details > div.container.h-mb-30 > div > div.col-lg-8.col-sm-12.col-xs-12.h-mb-15 > i').text
    tagline = tagline.replace('\n',' ')
  except:
    tagline = ''

    #-------------------------- ABOUT DATA CATCH --------------------------#
  try:
    driver.find_element_by_css_selector('#details > div.container.h-mb-30 > div > div.col-lg-8.col-sm-12.col-xs-12.h-mb-15 > p > a').click()
  except:
    print('No Read More Data ABOUT DATA :) :) :)')
  try:
    about = driver.find_element_by_css_selector('#details > div.container.h-mb-30 > div > div.col-lg-8.col-sm-12.col-xs-12.h-mb-15 > p > span:nth-child(1)').text
    about = about.replace('\n',' ')
  except:
    about = ''
    #---------------------------------- SERVICES EXTRACT------------------------------------#
  try:
    driver.find_element_by_css_selector('#details > div.container.container_xs.ng-scope > div > div > bdp-details-copy-points > div > div:nth-child(1) > div > div > div > ul > li:nth-child(19) > a > span:nth-child(3)').click() #view more button click
  except:
    print('LEFT POSTION No View More Found :) :) :)')
  try:
    driver.find_element_by_css_selector('#details > div.container.container_xs.ng-scope > div > div > bdp-details-copy-points > div > div:nth-child(2) > div > div > div > ul > li:nth-child(12) > a > span:nth-child(3)').click()
  except:
    print('RIGHT POSITION No View More Found :) :) :)')

  try: #https://www.truelocal.com.au/business/mitchell-cabinetmakers/noosaville 2x2::21
    driver.find_element_by_css_selector('#details > div.container.container_xs.ng-scope > div > div > bdp-details-copy-points > div > div:nth-child(2) > div > div > div > ul > li:nth-child(15) > a > span:nth-child(3)').click()
  except:
    print('2 x 2 :: 21 no more found :) :) ')

  try:
    driver.find_element_by_css_selector('#details > div.container.container_xs.ng-scope > div > div > bdp-details-copy-points > div > div:nth-child(2) > div:nth-child(1) > div > div > ul > li:nth-child(9) > a > span:nth-child(3)')
  except:
    print('MIDDLE POSTION No View More Found :) :) :)')

  try:
    driver.find_element_by_css_selector('#details > div.container.container_xs.ng-scope > div > div > bdp-details-copy-points > div > div:nth-child(1) > div:nth-child(2) > div > div > ul > li:nth-child(10) > a > span:nth-child(3)').click()
  except:
    print('SECOND LEFT POSITION No View More Found :) :) :)')


  #------------------------------------ SERVICES -------------------------------------------#
# left1=driver.find_element_by_css_selector('#details > div.container.container_xs.ng-scope > div > div > bdp-details-copy-points > div > div:nth-child(1) > div > h4').text
# right1=driver.find_element_by_css_selector('#details > div.container.container_xs.ng-scope > div > div > bdp-details-copy-points > div > div:nth-child(2) > div > h4').text
# middle1=driver.find_element_by_css_selector('#details > div.container.container_xs.ng-scope > div > div > bdp-details-copy-points > div > div:nth-child(2) > div:nth-child(1) > h4')
# left2=driver.find_element_by_css_selector('#details > div.container.container_xs.ng-scope > div > div > bdp-details-copy-points > div > div:nth-child(1) > div:nth-child(2) > h4').text
# middle2 =driver.find_element_by_css_selector('#details > div.container.container_xs.ng-scope > div > div > bdp-details-copy-points > div > div:nth-child(2) > div:nth-child(2) > h4').text
  try:
    #LEFT POSITION https://www.truelocal.com.au/business/advanced-craft-construction/kensington
    if(driver.find_element_by_css_selector('#details > div.container.container_xs.ng-scope > div > div > bdp-details-copy-points > div > div:nth-child(1) > div > h4').text=="Services"):
      services_selector='#details > div.container.container_xs.ng-scope > div > div > bdp-details-copy-points > div > div:nth-child(1) > div > div > div > ul'

    #RIGHT POSITION https://www.truelocal.com.au/business/solid-image-property-maintenance-and-home-improvements-pty-ltd/bentleigh-east
    elif(driver.find_element_by_css_selector('#details > div.container.container_xs.ng-scope > div > div > bdp-details-copy-points > div > div:nth-child(2) > div > h4').text=="Services"):
      services_selector='#details > div.container.container_xs.ng-scope > div > div > bdp-details-copy-points > div > div:nth-child(2) > div > div > div > ul'

    #RIGHT MIDDLE https://www.truelocal.com.au/business/chris-plastering/altona-meadows
    elif(driver.find_element_by_css_selector('#details > div.container.container_xs.ng-scope > div > div > bdp-details-copy-points > div > div:nth-child(2) > div:nth-child(1) > h4')=="Services"):
      services_selector='#details > div.container.container_xs.ng-scope > div > div > bdp-details-copy-points > div > div:nth-child(2) > div:nth-child(1) > div > div > ul'

    #SECOND ROW LEFT https://www.truelocal.com.au/business/great-aussie-patios/maddington
    elif(driver.find_element_by_css_selector('#details > div.container.container_xs.ng-scope > div > div > bdp-details-copy-points > div > div:nth-child(1) > div:nth-child(2) > h4').text=="Services"):
      services_selector = '#details > div.container.container_xs.ng-scope > div > div > bdp-details-copy-points > div > div:nth-child(1) > div:nth-child(2) > div > div > ul'

    #SECOND ROW MIDDLE  LIKE https://www.truelocal.com.au/business/pt-building-and-carpentry/carrum-downs *****HOTE PARE*******
    elif(driver.find_element_by_css_selector('#details > div.container.container_xs.ng-scope > div > div > bdp-details-copy-points > div > div:nth-child(2) > div:nth-child(2) > h4').text=="Services"):
      services_selector='#details > div.container.container_xs.ng-scope > div > div > bdp-details-copy-points > div > div:nth-child(2) > div:nth-child(2) > div > div > ul'

    #SECOND ROW LAST LIKE https://www.truelocal.com.au/business/pt-building-and-carpentry/carrum-downs  *****HOTE PARE ***
    else:
      services_selector='#details > div.container.container_xs.ng-scope > div > div > bdp-details-copy-points > div > div:nth-child(3) > div:nth-child(2) > div > div > ul'
  except:
    print('Services Not Found :) :) :) ')
  try:
    services = driver.find_element_by_css_selector(services_selector).text
    services = services.replace('\n',',')
    last=services.find(',View les') #Cut off view more from last
    services=services[:last]
  except:
    services = ''
    #------------------------ ABN NUMBER -----------------------------------#
  try:
    abn = driver.find_element_by_css_selector('#details > div.container.container_xs.ng-scope > div > div > div.col-lg-4.col-md-4.col-sm-6.col-xs-12 > div > div > div:nth-child(5) > bdp-details-abn-acn > div > span > span.text-frame').text
  except:
    abn = ''

   #---------------------------------WEBSITE-------------------------------#
  try:
    driver.find_element_by_css_selector('#details > div.container.container_xs.ng-scope > div > div > div.col-lg-4.col-md-4.col-sm-6.col-xs-12 > div > div > bdp-details-contact-website > a:nth-child(1) > span.text-frame.with-aditional-item > span').click()
    handles = driver.window_handles  
    driver.switch_to.window(handles[1]) 
    # print every open window page title 
    #print(driver.current_url)
    website = driver.current_url
  except:
    website = ''
  #------------------------------COMPANY NAME------------------------------#
  # try:
  #   if abn !="":
  #     abn_url=f'https://abr.business.gov.au/ABN/View?id={abn}'
  #     company=driver.find_element_by_xpath('/html/body/div[3]/div[3]/div/form/div/div[1]/table/tbody/tr[1]/td/span').text()
  # except:
  #   print('NO ABN, NO COMPANY NAME :) :) :) ')
  #--------------------------------GOOGLE MAP-----------------------------#

  sheet.cell(row=r,column=2).value=business
  sheet.cell(row=r,column=3).value=abn
  sheet.cell(row=r,column=4).value=services
  sheet.cell(row=r,column=5).value=address
  sheet.cell(row=r,column=6).value=phone
  sheet.cell(row=r,column=7).value=title
  sheet.cell(row=r,column=8).value=tagline
  sheet.cell(row=r,column=9).value=about
  # sheet.cell(row=r,column=10).value=gmap
  sheet.cell(row=r,column=11).value=website
  print("business : "+business)
  print("ABN : "+abn)
  print("services : "+services)
  print("Address : "+address)
  print("Phone : "+phone)
  print("Title : "+title)
  print("Tagline : "+tagline)
  print("About : "+about)
  # print("Gmap : "+gmap)
  print("website : "+website)
  wb.save('CarpenterData.xlsx')
  print("DATA INSERTED SUCCESSFULLY IN ROW "+str(r))
  driver.quit()